import sys
from openpyxl import load_workbook
import math

def is_number(x):
    if x is None:
        return False
    try:
        float(x)
        return True
    except:
        return False

def analyze(xlsx_path, sheet_name=None, max_start=40, look_forward=8):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet = wb[wb.sheetnames[0]]
    else:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return 2
        sheet = wb[sheet_name]

    rows = list(sheet.iter_rows(values_only=True))
    total_rows = len(rows)
    candidates = []
    for start in range(0, min(max_start, total_rows)):
        # examine next look_forward rows for numeric density
        end = min(start + look_forward, total_rows)
        numeric_counts = 0
        total_cells = 0
        for r in rows[start+1:end]:
            for c in r:
                total_cells += 1
                if is_number(c):
                    numeric_counts += 1
        density = numeric_counts / total_cells if total_cells>0 else 0
        candidates.append((start+1, density, numeric_counts, total_cells))

    # sort by density desc
    candidates.sort(key=lambda x: -x[1])
    print("Top candidate header positions (row, numeric_density, numeric_count, total_cells):")
    for pos, density, nc, tc in candidates[:10]:
        print(f"Row {pos}: density={density:.3f}, nums={nc}, total={tc}")
    return 0

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: detect_data_block.py <xlsx_path> [sheet_name]")
        sys.exit(2)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv)>2 else None
    sys.exit(analyze(path, sheet))
