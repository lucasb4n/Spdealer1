import sys
import re
from openpyxl import load_workbook

MONTH_PATTERNS = [
    r'jan', r'fev', r'mar', r'abr', r'mai', r'jun', r'jul', r'ago', r'set', r'out', r'nov', r'dez'
]

def score_row(row):
    score = 0
    matches = []
    for cell in row:
        if cell is None:
            continue
        text = str(cell).lower()
        for pat in MONTH_PATTERNS:
            if re.search(r'\b' + pat + r'\b', text):
                score += 1
                matches.append(text)
                break
    return score, matches

def main(xlsx_path, sheet_name=None, max_rows=50):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name is None:
        sheet = wb[wb.sheetnames[0]]
    else:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return 2
        sheet = wb[sheet_name]

    best = []
    for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if i > max_rows:
            break
        sc, matches = score_row(row)
        if sc > 0:
            best.append((i, sc, matches, row))

    if not best:
        print("No month-like header rows found in the first", max_rows, "rows.")
        return 0

    # sort by score desc then row index
    best.sort(key=lambda x: (-x[1], x[0]))
    print("Candidate header rows (top 10):")
    for idx, sc, matches, row in best[:10]:
        non_empty = [str(c) for c in row if c is not None]
        preview = " | ".join(non_empty[:10])
        print(f"Row {idx}: score={sc}, matches={matches}, preview={preview}")
    return 0

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: detect_month_header.py <xlsx_path> [sheet_name]")
        sys.exit(2)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    sys.exit(main(path, sheet))
