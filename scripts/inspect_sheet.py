#!/usr/bin/env python3
import sys
import openpyxl

def main():
    if len(sys.argv) < 2:
        print('Usage: inspect_sheet.py <xlsx> [sheet]')
        return 1
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv) > 2 else None
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet:
        if sheet not in wb.sheetnames:
            print('Sheet not found. Available:', wb.sheetnames)
            return 2
        ws = wb[sheet]
    else:
        ws = wb[wb.sheetnames[0]]
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=40, values_only=True), start=1):
        print(i, row)
    return 0

if __name__ == '__main__':
    raise SystemExit(main())
