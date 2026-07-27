import sys
from openpyxl import load_workbook

def preview(xlsx_path, sheet_name=None, start=1, end=60):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]] if sheet_name is None else wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    total = len(rows)
    end = min(end, total)
    for i in range(start-1, end):
        row = rows[i]
        nonempties = [(j+1, c) for j,c in enumerate(row) if c is not None and str(c).strip()!='']
        print(f"Row {i+1}: nonempty_count={len(nonempties)} -> {nonempties[:10]}")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: preview_rows.py <xlsx_path> [sheet_name] [start] [end]')
        sys.exit(2)
    path = sys.argv[1]
    sheet = sys.argv[2] if len(sys.argv)>2 else None
    start = int(sys.argv[3]) if len(sys.argv)>3 else 1
    end = int(sys.argv[4]) if len(sys.argv)>4 else 60
    preview(path, sheet, start, end)
